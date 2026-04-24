from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import EXCEL_COLUMNS
from .models import Tender


def _format_date(value) -> str:
    if isinstance(value, date):
        return value.isoformat()
    return ""


def _row_for(tender: Tender) -> list:
    return [
        tender.source,
        tender.title,
        tender.authority,
        _format_date(tender.publication_date),
        _format_date(tender.deadline),
        tender.estimated_value_eur if tender.estimated_value_eur is not None else "",
        ", ".join(tender.cpv_codes),
        (tender.scope or "")[:500],
        tender.reference,
        tender.url,
        ", ".join(tender.matched_keywords),
        tender.relevance_score,
    ]


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[column_letter].width = min(max(12, max_len + 2), 60)


def write_xlsx(tenders: List[Tender], path: Path) -> None:
    sorted_tenders = sorted(
        tenders,
        key=lambda t: (
            -(t.relevance_score or 0),
            -(t.publication_date.toordinal() if t.publication_date else 0),
        ),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Tenders"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row_idx, tender in enumerate(sorted_tenders, start=2):
        values = _row_for(tender)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        url_cell = ws.cell(row=row_idx, column=EXCEL_COLUMNS.index("URL") + 1)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    _autosize(ws)

    path = Path(path)
    try:
        wb.save(path)
    except PermissionError:
        alt = path.with_name(path.stem + "_1" + path.suffix)
        wb.save(alt)
        raise PermissionError(f"Original file locked; wrote to {alt} instead")
