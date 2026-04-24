"""Persistent storage - Excel workbook + snapshot files on disk.

The Excel workbook (``afas_monitor_data.xlsx``) is the single source of
truth. Snapshot HTML/text is too large to keep inline so it's written to
``snapshots/<url-hash>/<timestamp>.{html,txt}`` and referenced from the
workbook.
"""
from __future__ import annotations

import hashlib
import logging
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import app_base_dir, excel_path, snapshots_dir
from differ import DiffReport, PageFingerprint
from scraper import PageSnapshot

log = logging.getLogger(__name__)


SHEET_OVERVIEW = "Overzicht"
SHEET_PAGES = "Pagina's"
SHEET_CHANGES = "Wijzigingen"
SHEET_SNAPSHOTS = "Snapshots"

OVERVIEW_HEADERS = [
    "Datum (UTC)",
    "Pagina's gescand",
    "Nieuwe pagina's",
    "Verwijderde pagina's",
    "Tekst gewijzigd",
    "Structuur gewijzigd",
    "Totaal wijzigingen",
    "Rapport",
]

PAGES_HEADERS = [
    "URL",
    "Laatste status",
    "Laatst gecheckt (UTC)",
    "Text hash",
    "Structure hash",
    "Laatste snapshot (HTML)",
    "Laatste snapshot (TXT)",
]

CHANGES_HEADERS = [
    "Datum (UTC)",
    "URL",
    "Type wijziging",
    "Samenvatting",
]

SNAPSHOTS_HEADERS = [
    "URL",
    "Datum (UTC)",
    "HTTP-status",
    "Text hash",
    "Structure hash",
    "HTML-bestand",
    "Tekstbestand",
]

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")


@dataclass
class StoredPage:
    url: str
    last_status: int
    last_check: str
    text_hash: str
    structure_hash: str
    html_path: str
    text_path: str


def _url_dir_name(url: str) -> str:
    """Stable short directory name for a URL (filesystem-safe)."""
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:16]
    slug = re.sub(r"[^a-zA-Z0-9._-]+", "_", url)[:80].strip("_") or "page"
    return f"{slug}__{digest}"


def _timestamp_slug(iso_ts: str) -> str:
    return re.sub(r"[^0-9]+", "", iso_ts)[:14] or datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")


def _write_header(ws: Worksheet, headers: List[str]) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28


def _ensure_workbook(path: Path) -> Workbook:
    if path.is_file():
        try:
            wb = load_workbook(path)
        except Exception as exc:  # noqa: BLE001 - corrupt workbook
            log.warning("Existing workbook at %s is unreadable (%s) - recreating", path, exc)
            wb = Workbook()
            wb.remove(wb.active)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for sheet_name, headers in (
        (SHEET_OVERVIEW, OVERVIEW_HEADERS),
        (SHEET_PAGES, PAGES_HEADERS),
        (SHEET_CHANGES, CHANGES_HEADERS),
        (SHEET_SNAPSHOTS, SNAPSHOTS_HEADERS),
    ):
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _write_header(ws, headers)
    return wb


class Storage:
    """Thin wrapper around the Excel workbook and the snapshots folder."""

    def __init__(self, workbook_path: Optional[Path] = None, snap_dir: Optional[Path] = None) -> None:
        self.workbook_path = workbook_path or excel_path()
        self.snap_dir = snap_dir or snapshots_dir()

    # -- loading ---------------------------------------------------------

    def load_pages_index(self) -> Dict[str, StoredPage]:
        if not self.workbook_path.is_file():
            return {}
        try:
            wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            log.warning("Cannot open %s for reading: %s", self.workbook_path, exc)
            return {}
        out: Dict[str, StoredPage] = {}
        if SHEET_PAGES not in wb.sheetnames:
            wb.close()
            return out
        ws = wb[SHEET_PAGES]
        rows = ws.iter_rows(min_row=2, values_only=True)
        for row in rows:
            if not row or not row[0]:
                continue
            url = str(row[0])
            try:
                last_status = int(row[1]) if row[1] is not None else 0
            except (TypeError, ValueError):
                last_status = 0
            out[url] = StoredPage(
                url=url,
                last_status=last_status,
                last_check=str(row[2] or ""),
                text_hash=str(row[3] or ""),
                structure_hash=str(row[4] or ""),
                html_path=str(row[5] or ""),
                text_path=str(row[6] or ""),
            )
        wb.close()
        return out

    def load_previous_fingerprints(
        self, pages: Dict[str, StoredPage]
    ) -> Tuple[Dict[str, PageFingerprint], Dict[str, Tuple[str, str]]]:
        """Re-hydrate fingerprints + raw text/html for every known page."""
        from differ import compile_ignore_patterns, fingerprint  # local import to avoid cycles

        fingerprints: Dict[str, PageFingerprint] = {}
        raw: Dict[str, Tuple[str, str]] = {}
        base = app_base_dir()
        for url, page in pages.items():
            text = ""
            html = ""
            try:
                if page.text_path:
                    p = Path(page.text_path)
                    if not p.is_absolute():
                        p = base / p
                    if p.is_file():
                        text = p.read_text(encoding="utf-8", errors="replace")
                if page.html_path:
                    p = Path(page.html_path)
                    if not p.is_absolute():
                        p = base / p
                    if p.is_file():
                        html = p.read_text(encoding="utf-8", errors="replace")
            except OSError as exc:
                log.warning("Could not read snapshot for %s: %s", url, exc)
            # We use an empty pattern list here because snapshots on disk are
            # the raw captured content. The caller normalises via fingerprint().
            fp = fingerprint(url, text, html, page.last_check, compile_ignore_patterns([]))
            # But we DO trust the hashes stored in the sheet when they're present,
            # so comparisons stay consistent across normalisation rules.
            if page.text_hash:
                fp.text_hash = page.text_hash
            if page.structure_hash:
                fp.structure_hash = page.structure_hash
            fingerprints[url] = fp
            raw[url] = (text, html)
        return fingerprints, raw

    # -- writing ---------------------------------------------------------

    def write_snapshot_file(self, snap: PageSnapshot) -> Tuple[Path, Path]:
        """Persist HTML + text of one page and return (html_path, text_path)."""
        folder = self.snap_dir / _url_dir_name(snap.url)
        folder.mkdir(parents=True, exist_ok=True)
        ts = _timestamp_slug(snap.fetched_at)
        html_path = folder / f"{ts}.html"
        text_path = folder / f"{ts}.txt"
        try:
            html_path.write_text(snap.html or "", encoding="utf-8")
            text_path.write_text(snap.text or "", encoding="utf-8")
        except OSError as exc:
            log.warning("Failed to write snapshot files for %s: %s", snap.url, exc)
        return html_path, text_path

    def persist_run(
        self,
        snapshots: Iterable[PageSnapshot],
        fingerprints: Dict[str, PageFingerprint],
        report: DiffReport,
        report_file: Optional[Path] = None,
    ) -> None:
        """Write snapshots, update Pages, append to Changes/Snapshots/Overview."""
        run_date = datetime.now(timezone.utc).isoformat(timespec="seconds")
        base = app_base_dir()
        wb = _ensure_workbook(self.workbook_path)

        pages_ws = wb[SHEET_PAGES]
        snaps_ws = wb[SHEET_SNAPSHOTS]
        changes_ws = wb[SHEET_CHANGES]
        overview_ws = wb[SHEET_OVERVIEW]

        # Build a URL -> row index map for Pages so we can update in place.
        url_to_row: Dict[str, int] = {}
        for row_idx in range(2, pages_ws.max_row + 1):
            val = pages_ws.cell(row=row_idx, column=1).value
            if val:
                url_to_row[str(val)] = row_idx

        scanned = 0
        for snap in snapshots:
            scanned += 1
            if snap.error and snap.error in {"non-html", "stopped"}:
                continue
            if not snap.html and not snap.text:
                # Error case - still record status in Pages, no snapshot files.
                if snap.url not in url_to_row:
                    row = pages_ws.max_row + 1 if pages_ws.max_row >= 1 else 2
                    url_to_row[snap.url] = row
                r = url_to_row[snap.url]
                pages_ws.cell(row=r, column=1, value=snap.url)
                pages_ws.cell(row=r, column=2, value=snap.status or 0)
                pages_ws.cell(row=r, column=3, value=snap.fetched_at)
                continue

            html_path, text_path = self.write_snapshot_file(snap)
            fp = fingerprints.get(snap.url)
            text_hash = fp.text_hash if fp else ""
            struct_hash = fp.structure_hash if fp else ""

            rel_html = _rel_to(base, html_path)
            rel_text = _rel_to(base, text_path)

            snaps_ws.append([
                snap.url, snap.fetched_at, snap.status,
                text_hash, struct_hash, rel_html, rel_text,
            ])

            if snap.url in url_to_row:
                r = url_to_row[snap.url]
            else:
                r = (pages_ws.max_row + 1) if pages_ws.max_row >= 1 else 2
                url_to_row[snap.url] = r
            pages_ws.cell(row=r, column=1, value=snap.url)
            pages_ws.cell(row=r, column=2, value=snap.status)
            pages_ws.cell(row=r, column=3, value=snap.fetched_at)
            pages_ws.cell(row=r, column=4, value=text_hash)
            pages_ws.cell(row=r, column=5, value=struct_hash)
            pages_ws.cell(row=r, column=6, value=rel_html)
            pages_ws.cell(row=r, column=7, value=rel_text)

        for change in report.changes:
            changes_ws.append([
                run_date,
                change.url,
                change.change_type.value,
                change.summary,
            ])

        report_rel = _rel_to(base, report_file) if report_file else ""
        overview_ws.append([
            run_date,
            report.total_scanned or scanned,
            len(report.new_pages),
            len(report.removed_pages),
            len(report.text_changed),
            len(report.structure_changed),
            len(report.changes),
            report_rel,
        ])

        try:
            self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(self.workbook_path)
        except PermissionError as exc:
            # Excel might be holding the file open - surface a clear error.
            raise RuntimeError(
                f"Kan {self.workbook_path} niet opslaan (sluit Excel en probeer opnieuw): {exc}"
            ) from exc
        finally:
            wb.close()


def _rel_to(base: Path, path: Optional[Path]) -> str:
    if path is None:
        return ""
    try:
        return str(path.resolve().relative_to(base.resolve()))
    except ValueError:
        return str(path)
